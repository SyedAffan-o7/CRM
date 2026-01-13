from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
import csv
from io import TextIOWrapper
import openpyxl
from .models import Category, Subcategory
from .forms import CategoryForm, SubcategoryForm

@login_required
def products_home(request):
    return render(request, "Products/products_Home.html")

@login_required
def category_list(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.created_by = request.user
            category.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Category created successfully!',
                    'category': {
                        'id': category.id,
                        'name': category.name,
                        'description': category.description,
                        'created_at': category.created_at.strftime('%b %d, %Y')
                    }
                })
            else:
                messages.success(request, 'Category created successfully!')
                return redirect('products:category-list')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    
    elif request.method == 'PUT':
        data = json.loads(request.body)
        category_id = data.get('id')
        category = get_object_or_404(Category, id=category_id)
        
        category.name = data.get('name', category.name)
        category.description = data.get('description', category.description)
        category.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Category updated successfully!',
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'created_at': category.created_at.strftime('%b %d, %Y')
            }
        })
    
    elif request.method == 'DELETE':
        data = json.loads(request.body)
        category_id = data.get('id')
        category = get_object_or_404(Category, id=category_id)
        category.is_active = False
        category.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Category deleted successfully!'
        })
    
    else:
        form = CategoryForm()
    
    categories = Category.objects.filter(is_active=True)
    context = {
        'categories': categories,
        'form': form
    }
    return render(request, "Products/category_list.html", context)

@login_required
def subcategory_list(request):
    if request.method == 'POST':
        form = SubcategoryForm(request.POST)
        if form.is_valid():
            subcategory = form.save(commit=False)
            subcategory.created_by = request.user
            subcategory.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Subcategory created successfully!',
                    'subcategory': {
                        'id': subcategory.id,
                        'name': subcategory.name,
                        'category_name': subcategory.category.name,
                        'category_id': subcategory.category.id,
                        'description': subcategory.description,
                        'created_at': subcategory.created_at.strftime('%b %d, %Y')
                    }
                })
            else:
                messages.success(request, 'Subcategory created successfully!')
                return redirect('products:subcategory-list')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    
    elif request.method == 'PUT':
        data = json.loads(request.body)
        subcategory_id = data.get('id')
        subcategory = get_object_or_404(Subcategory, id=subcategory_id)
        
        subcategory.name = data.get('name', subcategory.name)
        subcategory.description = data.get('description', subcategory.description)
        if data.get('category_id'):
            subcategory.category_id = data.get('category_id')
        subcategory.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Subcategory updated successfully!',
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'category_name': subcategory.category.name,
                'category_id': subcategory.category.id,
                'description': subcategory.description,
                'created_at': subcategory.created_at.strftime('%b %d, %Y')
            }
        })
    
    elif request.method == 'DELETE':
        data = json.loads(request.body)
        # Support bulk delete if 'ids' list is provided
        ids = data.get('ids')
        if isinstance(ids, list) and ids:
            updated = Subcategory.objects.filter(id__in=ids).update(is_active=False)
            return JsonResponse({
                'success': True,
                'message': f'{updated} subcategory(s) deleted successfully!',
                'deleted_count': updated,
                'ids': ids,
            })
        # Fallback to single id delete
        subcategory_id = data.get('id')
        subcategory = get_object_or_404(Subcategory, id=subcategory_id)
        subcategory.is_active = False
        subcategory.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Subcategory deleted successfully!'
        })
    
    else:
        form = SubcategoryForm()
    
    subcategories = Subcategory.objects.filter(is_active=True).select_related('category')
    context = {
        'subcategories': subcategories,
        'form': form
    }
    return render(request, "Products/subcategory_list.html", context)

def product_list(request):
    return render(request, "Products/product_list.html")


@login_required
def import_categories_csv(request):
    """
    Import Categories/Subcategories from a CSV or XLSX file where:
    - Each COLUMN header is a Category name
    - Each non-empty CELL under a column is a Subcategory of that Category

    Example CSV:
        PPE,Electronics,Pharma
        Gloves,TV,Tablets
        Helmets,,Syringes
        Masks,Audio,

    Supports both .csv and .xlsx formats
    """
    if request.method == 'POST':
        upload = request.FILES.get('file')
        if not upload:
            messages.error(request, 'Please select a CSV or XLSX file to upload.')
            return redirect('products:products-import-csv')

        # Validate content type/basic extension
        filename = getattr(upload, 'name', '')
        is_csv = filename.lower().endswith('.csv')
        is_xlsx = filename.lower().endswith('.xlsx')

        if not (is_csv or is_xlsx):
            messages.error(request, 'Only .csv and .xlsx files are supported.')
            return redirect('products:products-import-csv')

        # Stats
        created_categories = 0
        reactivated_categories = 0
        created_subcategories = 0
        reactivated_subcategories = 0
        skipped_blank = 0

        try:
            data_rows = []

            if is_csv:
                # Handle CSV files
                wrapper = TextIOWrapper(upload.file, encoding='utf-8-sig')
                reader = csv.DictReader(wrapper)

                # Handle case: empty file
                if not reader.fieldnames:
                    messages.error(request, 'CSV seems empty or malformed (no headers).')
                    return redirect('products:products-import-csv')

                # Convert to list of dicts for consistent processing
                data_rows = list(reader)
                headers = reader.fieldnames

            elif is_xlsx:
                # Handle XLSX files
                workbook = openpyxl.load_workbook(upload.file)
                worksheet = workbook.active

                # Get headers from first row
                headers = []
                for col in range(1, worksheet.max_column + 1):
                    header_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(header_value) if header_value else f'Column_{col}')

                # Get data rows
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value else ''
                    data_rows.append(row_data)

            # Handle case: empty file
            if not headers:
                messages.error(request, f'{filename.split(".")[-1].upper()} seems empty or malformed (no headers).')
                return redirect('products:products-import-csv')

            # Map/ensure all categories from headers exist first
            category_cache = {}
            for header in headers:
                if header is None:
                    continue
                cat_name = str(header).strip()
                if not cat_name:
                    continue

                category, created = Category.objects.get_or_create(
                    name=cat_name,
                    defaults={
                        'description': '',
                        'created_by': request.user,
                        'is_active': True,
                    }
                )
                if created:
                    created_categories += 1
                else:
                    # If category exists but was inactive, reactivate
                    if not category.is_active:
                        category.is_active = True
                        category.save(update_fields=['is_active'])
                        reactivated_categories += 1
                category_cache[cat_name] = category

            # Iterate rows: each cell under a header is a subcategory
            for row in data_rows:
                for header, value in row.items():
                    if header is None:
                        continue
                    cat_name = str(header).strip()
                    if not cat_name:
                        continue
                    sub_name = (str(value).strip() if value is not None else '')
                    if not sub_name:
                        skipped_blank += 1
                        continue

                    category = category_cache.get(cat_name)
                    if not category:
                        # Safety: skip if header wasn't processed properly
                        continue

                    subcat, created = Subcategory.objects.get_or_create(
                        category=category,
                        name=sub_name,
                        defaults={
                            'description': '',
                            'created_by': request.user,
                            'is_active': True,
                        }
                    )
                    if created:
                        created_subcategories += 1
                    else:
                        if not subcat.is_active:
                            subcat.is_active = True
                            subcat.save(update_fields=['is_active'])
                            reactivated_subcategories += 1

            messages.success(
                request,
                (
                    f"Import completed from {filename.split('.')[-1].upper()} file. "
                    f"Categories created: {created_categories}, reactivated: {reactivated_categories}. "
                    f"Subcategories created: {created_subcategories}, reactivated: {reactivated_subcategories}. "
                    f"Skipped blank cells: {skipped_blank}."
                )
            )
            return redirect('products:category-list')

        except Exception as e:
            messages.error(request, f'Failed to import {filename.split(".")[-1].upper()}: {e}')
            return redirect('products:products-import-csv')

    # GET: show upload form with instructions
    return render(request, 'Products/category_import.html', {})


def download_example_csv(request):
    """
    Generate and download a random example CSV file with different categories and subcategories each time.
    """
    import csv
    import random
    from io import StringIO
    from django.http import HttpResponse

    # Define pools of categories and subcategories for random generation
    category_templates = [
        'Electronics', 'Clothing', 'Home & Garden', 'Sports & Outdoors', 'Books',
        'Beauty & Personal Care', 'Automotive', 'Toys & Games', 'Health & Wellness',
        'Office Supplies', 'Pet Supplies', 'Baby Products', 'Industrial Equipment',
        'Food & Beverages', 'Art & Crafts', 'Musical Instruments'
    ]

    subcategory_templates = {
        'Electronics': ['Smartphones', 'Laptops', 'Tablets', 'Headphones', 'Speakers', 'Cameras', 'TVs', 'Gaming Consoles', 'Smart Watches', 'Chargers'],
        'Clothing': ['Shirts', 'Pants', 'Dresses', 'Shoes', 'Jackets', 'Underwear', 'Socks', 'Hats', 'Belts', 'Accessories'],
        'Home & Garden': ['Furniture', 'Decor', 'Kitchenware', 'Bedding', 'Bath', 'Storage', 'Tools', 'Plants', 'Lighting', 'Cleaning Supplies'],
        'Sports & Outdoors': ['Exercise Equipment', 'Outdoor Gear', 'Sports Apparel', 'Bicycles', 'Camping', 'Fishing', 'Team Sports', 'Water Sports', 'Winter Sports', 'Fitness Accessories'],
        'Books': ['Fiction', 'Non-Fiction', 'Textbooks', 'Children\'s Books', 'Biographies', 'Cookbooks', 'Travel Guides', 'Technical Books', 'Comics', 'Magazines'],
        'Beauty & Personal Care': ['Skincare', 'Makeup', 'Hair Care', 'Fragrances', 'Nail Care', 'Bath & Body', 'Men\'s Grooming', 'Tools & Accessories', 'Wellness Products', 'Spa Products'],
        'Automotive': ['Car Parts', 'Accessories', 'Tools', 'Electronics', 'Interior', 'Exterior', 'Maintenance', 'Safety', 'Performance', 'Cleaning'],
        'Toys & Games': ['Action Figures', 'Dolls', 'Building Sets', 'Board Games', 'Puzzles', 'Outdoor Toys', 'Educational Toys', 'Electronic Toys', 'Stuffed Animals', 'Arts & Crafts'],
        'Health & Wellness': ['Vitamins', 'Supplements', 'Fitness Equipment', 'Personal Care', 'Medical Supplies', 'Herbal Products', 'Essential Oils', 'Wellness Devices', 'Books & Media', 'Testing Kits'],
        'Office Supplies': ['Writing Instruments', 'Paper Products', 'Filing', 'Desk Accessories', 'Electronics', 'Furniture', 'Breakroom Supplies', 'Mailing Supplies', 'Presentation Supplies', 'Art Supplies'],
        'Pet Supplies': ['Food', 'Toys', 'Bedding', 'Grooming', 'Health Care', 'Accessories', 'Training', 'Travel', 'Aquarium', 'Bird Supplies'],
        'Baby Products': ['Diapers', 'Clothing', 'Feeding', 'Bath & Potty', 'Nursery', 'Safety', 'Toys', 'Car Seats', 'Strollers', 'Health & Baby Care'],
        'Industrial Equipment': ['Tools', 'Safety Equipment', 'Hardware', 'Electrical', 'Plumbing', 'HVAC', 'Cleaning', 'Storage', 'Material Handling', 'Test Equipment'],
        'Food & Beverages': ['Fruits', 'Vegetables', 'Meat & Poultry', 'Dairy', 'Bakery', 'Beverages', 'Snacks', 'Condiments', 'Canned Goods', 'Frozen Foods'],
        'Art & Crafts': ['Paints', 'Brushes', 'Canvas', 'Paper', 'Markers', 'Crayons', 'Clay', 'Beads', 'Yarn', 'Craft Kits'],
        'Musical Instruments': ['Guitars', 'Keyboards', 'Drums', 'Amplifiers', 'Microphones', 'Speakers', 'Accessories', 'Sheet Music', 'Instrument Care', 'Recording Equipment']
    }

    # Randomly select 3-5 categories for this example
    num_categories = random.randint(3, 5)
    selected_categories = random.sample(category_templates, num_categories)

    # Generate CSV content
    output = StringIO()
    writer = csv.writer(output)

    # Create headers (category names)
    writer.writerow(selected_categories)

    # Generate 4-6 rows of subcategories with some blank cells for realism
    num_rows = random.randint(4, 6)
    for row_num in range(num_rows):
        row_data = []
        for category in selected_categories:
            # 70% chance of having a subcategory in each cell
            if random.random() < 0.7:
                subcategories = subcategory_templates.get(category, ['Item'])
                subcategory = random.choice(subcategories)
                row_data.append(subcategory)
            else:
                row_data.append('')  # Blank cell

        writer.writerow(row_data)

    # Create HTTP response with CSV content
    csv_content = output.getvalue()
    output.close()

    response = HttpResponse(csv_content, content_type='text/csv')
    filename = f'example_categories_{random.randint(1000, 9999)}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
